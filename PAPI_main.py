#9h20min

#TO DO
#Calcular IR
#Criar e Formatar Dashboard (INICIO - COM PORCENTAGEM)


import os
import requests             #Requisição HTTP
import win32com.client      #Trabalhar com MacroVBA

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter                                #Biblioteca para Função de Converte para Letras Índice da Coluna           
from openpyxl.styles import Alignment, Font, Border, Side          #Biblioteca Para Formatação das Células
from datetime import datetime                                               #Biblioteca Converte string para datetime
#Alignment - Alinhamento
#Font - Font
#Border e Side- Borda


#Variavel Global
borda_padrao = Side(border_style="thin", color="000000")
borda_externa = Side(border_style="medium", color="000000")

##### FUNÇÕES PARA CRIAÇÃO E FORMATAÇÃO DE ESTILO #####

#Criar Novo Arquivo e Formatar Abas Padrão
def criar_arquivo(arquivo):
    aba_inicio = arquivo.active                             #Seleciona Aba Ativa - Criada Automaticamente
    aba_inicio.title = "DASHBOARD"                             #Nomeia Aba Criada Automaticamente
    ocultar_grades(aba_inicio)
    
    
    nova_aba = arquivo.create_sheet(title="TOTAL_APORTES")  #Nova Aba para Conter Todos os Aportes
    ocultar_grades(nova_aba)

    #Alterar Largura das Colunas
    nova_aba.column_dimensions['A'].width = 8.8
    nova_aba.column_dimensions['B'].width = 8.8
    alterar_largura_colunas_sequencia(nova_aba, 2, 6, 12.8)

    alterar_altura_linha(nova_aba, 1)
    nova_aba.freeze_panes = "A2"    #Congelar Cabeçalho

    #Preencher Cabeçalho da Planilha
    nova_aba.cell(row=1, column=1).value = "Ativos"
    nova_aba.cell(row=1, column=2).value = "Cotas"
    nova_aba.cell(row=1, column=3).value = "Preço"
    nova_aba.cell(row=1, column=4).value = "Investido"
    nova_aba.cell(row=1, column=5).value = "Data"

    #Alterção do Estilo e Borda da Célula Padrão
    for celula in nova_aba[1]:
        alterar_estilo_celula(celula)
        alterar_borda(celula, borda_padrao, borda_padrao, borda_padrao, borda_padrao)


#Criar Aba Nova para FII e Formatar
def criar_aba(arquivo, nome_ativo):
    nova_aba = arquivo.create_sheet(title=nome_ativo)       #Cria Nova Aba

    ocultar_grades(nova_aba)

    #Preencher Cabeçalho da Planilha
    nova_aba.merge_cells('B2:E2')                           #Mesclar Celulas
    nova_aba.cell(row=2, column=2).value = nome_ativo
    nova_aba.cell(row=3, column=2).value = "Cotas"
    nova_aba.cell(row=3, column=3).value = "Preço"
    nova_aba.cell(row=3, column=4).value = "Investido"
    nova_aba.cell(row=3, column=5).value = "Data"

    nova_aba.merge_cells('G2:J2')
    nova_aba.cell(row=2, column=7).value = "IR"
    nova_aba.cell(row=3, column=7).value = "Cotas"
    nova_aba.cell(row=3, column=8).value = "Preço Médio"
    nova_aba.cell(row=3, column=9).value = "Total"
    nova_aba.cell(row=3, column=10).value = "Ano"

    nova_aba.cell(row=2, column=12).value = "Total Cotas"
    nova_aba.cell(row=2, column=13).value = "Preço Atual"
    nova_aba.cell(row=2, column=14).value = "Total"
    nova_aba.cell(row=3, column=13).value = cotacao_atual(nome_ativo)
    nova_aba.cell(row=3, column = 13).number_format = 'R$ #,##0.00'

    nova_aba.freeze_panes = "A4"    #Congelar Cabeçalho

    #Alterar Altura das Linhas 1 a 3
    for linha in range(1, 4):
        alterar_altura_linha(nova_aba, linha)

    #Alterar Largura das Colunas
    nova_aba.column_dimensions['A'].width = 2.8
    nova_aba.column_dimensions['B'].width = 8.8
    nova_aba.column_dimensions['F'].width = 4.8
    nova_aba.column_dimensions['K'].width = 4.8

    alterar_largura_colunas_sequencia(nova_aba, 2, 6, 12.8)
    alterar_largura_colunas_sequencia(nova_aba, 7, 11, 14.8)
    alterar_largura_colunas_sequencia(nova_aba, 12, 15, 14.8)
    
    #Alterção do Estilo da Célula em Sequência com o Mesmo Padrão
    for linha in range(1,4):
        for celula in nova_aba[linha]:
            alterar_estilo_celula(celula)
    
    for linha in range(7,11):
        for celula in nova_aba[linha]:
            alterar_estilo_celula(celula)

    #Alterção da Borda da Célula Mesclada Para Titulo
    for coluna in range(2, 6):
        celula = nova_aba.cell(row=2, column=coluna)
        alterar_borda(celula, borda_externa, borda_externa, borda_externa, borda_externa)
    for coluna in range(7, 11):
        celula = nova_aba.cell(row=2, column=coluna)
        alterar_borda(celula, borda_externa, borda_externa, borda_externa, borda_externa)
    for coluna in range(12,15):
        for linha in range(2,4):
            celula = nova_aba.cell(row=linha, column=coluna)
            alterar_borda(celula, borda_externa, borda_externa, borda_externa, borda_externa)

    #Alterção da Borda da Célula Para SubTitulo
    alterar_borda(nova_aba.cell(row=3, column=2), borda_externa, borda_padrao, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=3), borda_padrao, borda_padrao, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=4), borda_padrao, borda_padrao, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=5), borda_padrao, borda_externa, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=7), borda_externa, borda_padrao, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=8), borda_padrao, borda_padrao, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=9), borda_padrao, borda_padrao, borda_externa, borda_externa)
    alterar_borda(nova_aba.cell(row=3, column=10), borda_padrao, borda_externa, borda_externa, borda_externa)


#Oculta as Linhas de Grade da Aba
def ocultar_grades(aba):
    aba.sheet_view.showGridLines = False


#Alterar Altura da Linha
def alterar_altura_linha(aba, linha):
    aba.row_dimensions[linha].height = 14


#Alterar Largura de Colunas em Sequência com o Mesmo Valor - Reutilização de Código
def alterar_largura_colunas_sequencia(nova_aba, coluna_inicio, coluna_final, tamanho):
    for coluna in range(coluna_inicio, coluna_final):  
        letra_coluna = get_column_letter(coluna)                    #Converte para Letras
        nova_aba.column_dimensions[letra_coluna].width = tamanho    #Alterar Largura de Coluna


#API para obter Cotação Em Tempo Real
def cotacao_atual(aba):
    url_api = "http://b3api.me/api/quote/"+aba
    try:
        resposta = requests.get(url_api)    #Get HTTP
        resposta.raise_for_status()         #Garante que só respostas 200 (OK) vai continuar o código
        cota = resposta.json()              #Dicionario em JSON
        return cota["price"]                #Retorna Cota
    except requests.exceptions.RequestException as erro:    #Exceção de Erro HTTP
        print(f'Erro ao Conectar com API: {erro}')
        return None


#Adicionar MacroVBA
def macro_vba(nome_arquivo):
    excel = win32com.client.Dispatch("Excel.Application")   #Abrir o Excel via COM
    #Component Object Model.
    #Microsoft que permite que programas diferentes conversem entre si e controlem uns aos outros
    excel.Visible = False  # Não exibe a janela do Excel

    arquivo = excel.Workbooks.Open("Ativos B3.xlsm")

    # Adiciona macro no módulo do Ativo
    #Requisição HTTP quando abre o arquivo
    vba_code = '''
    Private Sub Workbook_Open()
        Dim aba As Worksheet
        Dim http As Object
        Dim url As String
        Dim resposta As String
        Dim preco As Double
        Dim inicio As Long, fim As Long
        Dim codigo As String

        For Each aba In ThisWorkbook.Worksheets
            If aba.Name <> "DASHBOARD" And aba.Name <> "TOTAL_APORTES" Then
                Set http = CreateObject("MSXML2.XMLHTTP")
                codigo = aba.Name
                url = "http://b3api.me/api/" & codigo

                On Error Resume Next
                http.Open "GET", url, False
                http.Send

                If http.Status = 200 Then
                    resposta = http.responseText

                    ' Extrair o valor da chave "price" do JSON (simplesmente com string, sem usar parser)
                    inicio = InStr(resposta, """price"":") + Len("""price"":")
                    fim = InStr(inicio, resposta, ",")
                    If inicio > 0 And fim > inicio Then
                        preco = Val(Mid(resposta, inicio, fim - inicio))
                        aba.Range("M3").Value = preco
                    End If
                End If

                On Error GoTo 0
            End If
        Next aba
    End Sub
    '''

    # Acessar o módulo "ThisWorkbook" e adicionar VBA
    arquivo.VBProject.VBComponents("ThisWorkbook").CodeModule.AddFromString(vba_code)
    arquivo.Save()
    arquivo.Close()     #Fechar Arquivo
    excel.Quit()        #Fechar Excel
    print("Macro com Sucesso")


#Alterar Estilo da Célula com o Mesmo Padrão - Reutilização de Código
# Para Criação e Adição??
def alterar_estilo_celula(celula):
    celula.alignment = Alignment(horizontal='center', vertical='center')  #Centralizar Texto
    celula.font = Font(name="Arial", size=12, bold=False, italic=False, color="000000")
    #Nome da Fonte; Tamanho; Negrito; Italico; Cor


#Bordas
def alterar_borda(celula, left, right, top, bottom):
    celula.border = Border(left=left, right=right, top=top, bottom=bottom)


##### FUNÇÕES PARA ALIMENTAR TABELAS #####

#Obter Linha Vazia
#Vai até ultima que sofreu modificação (max_row)
def ultima_linha(aba):
    #Pula Primeira Linha (Vai ser Cabeçalho ou Vazia na Aba do Ativo)
    #For: max_row+3 para ter linhas vazias extras
    contador = 0
    for row in range(2, aba.max_row + 3):
        if aba.cell(row=row, column=2).value is None:
            contador += 1
        else:
            contador = 0
        if contador == 2: 
            return row-1
    return aba.max_row
    #Duas linhas Vazias Consecutivas Retorna Linha Anterior (Primeira Linha Vazia na Sequencia)


#Função para Adicionar Dados na Aba       
def adicionar_aporte(planilha, nome_fundo, cotas, valor, data, borda, aba_selecionada):
    #Seleciona Aba Especifica do Ativo ou Aba Geral
    aba = planilha[aba_selecionada]

    #Obter Linha
    linha = ultima_linha(aba)

    #Verificar se é Primeiro Input ou Não e Se Linha Anterior Não É Vazia
    if linha <= 4 and aba_selecionada != "TOTAL_APORTES":
        ano_anterior = "01/01/0001"       #Adiciona um Valor Para Comparar Porque é Primeiro Input (Aba Ativo)
        ano_anterior = datetime.strptime(ano_anterior, "%d/%m/%Y")
    elif linha <= 2 and aba_selecionada == "TOTAL_APORTES":
        ano_anterior = "01/01/0001"       #Adiciona um Valor Para Comparar Porque é Primeiro Input (Aba Total Aportes)
        ano_anterior = datetime.strptime(ano_anterior, "%d/%m/%Y")
    elif aba.cell(row=linha-1, column=2).value is not None:
        #Como Pula então talvez depois ele pega a linha vazia
        #Se conteudo tiver vazio
        linha_anterior = linha - 1
        ano_anterior = aba.cell(row=linha_anterior, column = 5).value       #Pega Ultimo Ano Digitado formato datetime
    else:
        linha_anterior = linha
        ano_anterior = aba.cell(row=linha_anterior, column = 5).value       #Pega Ultimo Ano Digitado formato datetime'''

    # Se Não é Aba Geral e Se É Ano Diferente (Se Ano Digitado é Diferente do Ultimo Adicionado) - Pular uma Linha
    if aba_selecionada != "TOTAL_APORTES" and ano_anterior.year != data.year and linha > 4:
        linha += 1
        borda = borda_externa
        borda_topo = borda_externa
        #Se Ano Anterior Foi Apenas um Aporte - Estilo
        if aba.cell(row=linha_anterior-1, column = 2).value is None:
            borda_top = borda_externa
        else:
            borda_top = borda_padrao
        #Formatar Borda do Fim do Ano Anterior
        alterar_borda(aba.cell(row=linha_anterior, column = 2),borda, borda_padrao, borda_top, borda_externa)
        alterar_borda(aba.cell(row=linha_anterior, column = 3),borda_padrao, borda_padrao, borda_top, borda_externa)
        alterar_borda(aba.cell(row=linha_anterior, column = 4),borda_padrao, borda_padrao, borda_top, borda_externa)
        alterar_borda(aba.cell(row=linha_anterior, column = 5),borda_padrao, borda, borda_top, borda_externa)    
    else:
        borda_topo = borda_padrao

    #Adicionar Conteudo na Planilha
    if aba_selecionada == "TOTAL_APORTES":
        aba.cell(row=linha, column = 1).value = nome_fundo
        alterar_borda(aba.cell(row=linha, column = 1),borda, borda_padrao, borda_padrao, borda_padrao)
        alterar_estilo_celula(aba.cell(row=linha, column = 1))
    aba.cell(row=linha, column = 2).value = cotas
    aba.cell(row=linha, column = 3).value = valor
    aba.cell(row=linha, column = 3).number_format = 'R$ #,##0.00'
    aba.cell(row=linha, column = 4).value = cotas * valor
    aba.cell(row=linha, column = 4).number_format = 'R$ #,##0.00'
    aba.cell(row=linha, column = 5).value = data
    aba.cell(row=linha, column = 5).number_format = "DD/MM/YYYY"

    #Formatar Estilo após Input
    for coluna in range(2, 6):
        alterar_estilo_celula(aba.cell(row=linha, column = coluna))
        alterar_altura_linha(aba, linha)
    alterar_borda(aba.cell(row=linha, column = 2),borda, borda_padrao, borda_topo, borda_padrao)
    alterar_borda(aba.cell(row=linha, column = 3),borda_padrao, borda_padrao, borda_topo, borda_padrao)
    alterar_borda(aba.cell(row=linha, column = 4),borda_padrao, borda_padrao, borda_topo, borda_padrao)
    alterar_borda(aba.cell(row=linha, column = 5),borda_padrao, borda, borda_topo, borda_padrao)    


######################################################

caminho = os.path.dirname(os.path.abspath(__file__))                        #Pega Caminho de Onde está o Arquivo
os.chdir(caminho)                                                           #Alterar Diretório de Execução do Script Python
#Garantir que o Python esteja da pasta do projeto 


nome_arquivo = "Ativos B3.xlsx"

#Verificação se Arquivo Existe
if os.path.exists(nome_arquivo):
    arquivo = load_workbook(nome_arquivo) #Carregar Arquivo - sem Macro
    # Como tá na mesma página, não precisa indicar caminho'''
else:
   arquivo = Workbook()                    #Cria Novo Arquivo
   criar_arquivo(arquivo)

nome_ativo = input("ATIVO: ")
nome_ativo = nome_ativo.upper()             #Converter Toda String Para Maiusculo
cotas = 10
valor = 12
data = "25/12/2020"
data = datetime.strptime(data, "%d/%m/%Y")  # Converte String para datetime
if nome_ativo not in arquivo.sheetnames:    #Se Ativo Não Tem Aba
    criar_aba(arquivo, nome_ativo)
adicionar_aporte(arquivo, nome_ativo, cotas, valor, data, borda_externa, nome_ativo)
adicionar_aporte(arquivo, nome_ativo, cotas, valor, data, borda_padrao, "TOTAL_APORTES")

arquivo.save(nome_arquivo)    #Sobrescreve o Arquivo/ Outro Nome Gera Outro Arquivo